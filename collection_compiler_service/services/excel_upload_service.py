import asyncio
import datetime as dt
import io
import json
import logging
import re
import zipfile
from decimal import Decimal, InvalidOperation
from typing import Any
from urllib.parse import quote
from xml.etree import ElementTree as ET
import openpyxl
from fastapi import HTTPException
from collection_compiler_service.model.FileUploadRecord import FileUploadRecord
from collection_compiler_service.repository.CaseAssignmentDetailRepository import CaseAssignmentDetailRepository
from collection_compiler_service.repository.FileUploadRecordRepository import FileUploadRecordRepository
from collection_compiler_service.schemas.CaseAssignmentDetailSchema import CaseAssignmentDetailPatchRequest
from common.model.BaseDBModel import generate_snowflake_id
from common.schemas.CommonResult import Result


logger = logging.getLogger("collection_compiler_service")


class ExcelUploadService:
    FIELD_MAPPING = {
        "案件ID": "case_id",
        "姓名": "debtor_name",
        "UID": "uid",
        "户籍地": "household_address",
        "进件编码": "application_code",
        "债转公告链接": "transfer_notice_url",
        "债转公告序号": "transfer_notice_no",
        "债转协议序号": "transfer_agreement_no",
        "项目名称": "project_name",
        "省": "province",
        "市": "city",
        "委案日期": "entrust_date",
        "委案批次号": "entrust_batch_no",
        "委案机构": "entrust_org",
        "委案剩本金额": "entrusted_principal_balance",
        "委案债权总额": "entrusted_total_claim",
        "处置方式": "disposal_type",
    }
    REQUIRED_COLUMNS = list(FIELD_MAPPING.keys())
    CRITICAL_FIELDS = {"UID", "进件编码"}
    IGNORED_MISSING_FIELDS = {"姓名", "户籍地", "省", "市"}
    NS = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    async def process_excel_bytes(self, file_bytes: bytes, filename: str, file_hash: str | None = None) -> Result:
        if not file_bytes:
            return Result.fail("上传文件为空", code=400)
        if not filename.lower().endswith(".xlsx"):
            return Result.fail("目前仅支持 .xlsx 文件", code=400)

        if file_hash:
            repo = FileUploadRecordRepository(db_name="collection_compiler")
            async with repo as repository:
                existing = await repository.find_by_hash(file_hash)
            if existing:
                logger.info("Excel 秒传命中: hash=%s", file_hash)
                return Result.success({"message": "文件已存在，跳过处理", "filename": filename, "file_hash": file_hash, "instant": True})

        try:
            raw_rows = self.parse_xlsx_rows(file_bytes)
        except HTTPException as e:
            return Result.fail(e.detail if isinstance(e.detail, str) else str(e.detail), code=e.status_code)
        row_contexts = [self.build_row_context(row) for row in raw_rows]
        validation_issues, validation_summary = self.build_validation_issues(row_contexts)

        issue_row_nums = {item["row_num"] for item in validation_issues}
        valid_records = [item["record"] for item in row_contexts if item["row_num"] not in issue_row_nums]
        defect_records = [item["record"] for item in row_contexts if item["row_num"] in issue_row_nums]

        # 按行归集最高严重级别：critical > warning
        row_severity: dict[int, str] = {}
        for issue in validation_issues:
            row_num = issue["row_num"]
            if row_severity.get(row_num) != "critical":
                row_severity[row_num] = issue["severity"]

        severity_to_status = {"critical": -2, "warning": -1}
        for item in row_contexts:
            item["record"]["case_status"] = severity_to_status.get(row_severity.get(item["row_num"]), 0)
            if file_hash:
                item["record"]["file_hash"] = file_hash

        logger.info("========== Excel处理完成 ==========")
        logger.info("文件名: %s", filename)
        logger.info("数据行数: %s", len(row_contexts))
        logger.info("严重错误: %s, 一般警告: %s", validation_summary["critical_count"], validation_summary["warning_count"])
        logger.info("有效记录数量: %s", len(valid_records))
        for index, item in enumerate(row_contexts, 1):
            logger.info("第%s行: %s", index, json.dumps(item["record"], ensure_ascii=False))
        logger.info("========== Excel处理完成 ==========")

        # 有缺失的数据同步入库，等待完成后再返回前端
        if defect_records:
            asyncio.create_task(self._insert_records(defect_records, filename))

        # 完整数据异步入库，不阻塞响应
        if valid_records:
            asyncio.create_task(self._insert_records(valid_records, filename))

        # 写入 hash 记录
        if file_hash:
            repo = FileUploadRecordRepository(db_name="collection_compiler")
            async with repo as repository:
                existing = await repository.find_by_hash(file_hash)
                if not existing:
                    entity = FileUploadRecord(
                        file_hash=file_hash,
                        filename=filename,
                        file_path=filename,
                        file_size=len(file_bytes),
                        file_type="excel",
                    )
                    await repository.save(entity)

        return Result.success({
            "message": "Excel 处理完成",
            "filename": filename,
            "total": len(row_contexts),
            "preview": [item["record"] for item in row_contexts[:3]],
            "validation_summary": validation_summary,
            "validation_issues": validation_issues[:200],
            "insert_summary": {
                "candidate_count": len(defect_records) + len(valid_records),
            },
        })

    async def _insert_records(self, valid_records: list[dict], filename: str) -> None:
        try:
            repo = CaseAssignmentDetailRepository(db_name="collection_compiler")
            async with repo as repository:
                inserted = await repository.save_batch_data(valid_records)
            logger.info("后台插入完成: %s 条, 文件: %s", len(inserted), filename)
        except Exception as e:
            logger.error("后台插入失败: %s, 文件: %s", e, filename)

    async def patch_and_validate(self, req: CaseAssignmentDetailPatchRequest) -> dict[str, Any]:
        """接收用户修改后的数据，重新校验并更新入库"""
        updates = req.model_dump(exclude={"id"}, exclude_none=True)

        # 重新校验：检查非忽略字段是否仍有缺失
        non_ignored = {
            field_label: field_key
            for field_label, field_key in self.FIELD_MAPPING.items()
            if field_label not in self.IGNORED_MISSING_FIELDS
        }
        missing_critical = []
        missing_warning = []
        for field_label, field_key in non_ignored.items():
            if not updates.get(field_key):
                if field_label in self.CRITICAL_FIELDS:
                    missing_critical.append(field_label)
                else:
                    missing_warning.append(field_label)

        if missing_critical:
            new_status = -2
        elif missing_warning:
            new_status = -1
        else:
            new_status = 0

        updates["case_status"] = new_status

        repo = CaseAssignmentDetailRepository(db_name="collection_compiler")
        async with repo as repository:
            await repository.update_by_id_selective(req.id, updates)

        return {
            "id": str(req.id),
            "case_status": new_status,
            "missing_critical": missing_critical,
            "missing_warning": missing_warning,
            "message": "数据已更新" if new_status == 0 else "数据已更新，但仍有字段缺失",
        }

    def parse_xlsx_rows(self, file_bytes: bytes) -> list[dict[str, Any]]:
        zf = zipfile.ZipFile(io.BytesIO(file_bytes))

        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall("a:si", self.NS):
                texts = [t.text or "" for t in si.findall(".//a:t", self.NS)]
                shared_strings.append("".join(texts))

        wb = ET.fromstring(zf.read("xl/workbook.xml"))
        first_sheet = wb.find("a:sheets/a:sheet", self.NS)
        if first_sheet is None:
            raise HTTPException(status_code=400, detail="Excel 文件无工作表")

        sheet_rid = first_sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        target = None
        for rel in rels.findall("pr:Relationship", self.NS):
            if rel.attrib.get("Id") == sheet_rid:
                target = rel.attrib.get("Target")
                break
        if not target:
            raise HTTPException(status_code=400, detail="无法获取工作表路径")

        sheet_path = target if target.startswith("xl/") else f"xl/{target}"
        sheet_root = ET.fromstring(zf.read(sheet_path))

        parsed_rows: list[tuple[int, list[Any]]] = []
        for row in sheet_root.findall("a:sheetData/a:row", self.NS):
            row_num = int(row.attrib.get("r", "0"))
            cells: dict[int, Any] = {}
            for cell in row.findall("a:c", self.NS):
                ref = cell.attrib.get("r", "")
                match = re.match(r"([A-Z]+)(\d+)", ref)
                if not match:
                    continue
                col_num = self.col_to_num(match.group(1))
                cells[col_num] = self.get_cell_value(cell, shared_strings)

            if cells:
                max_col = max(cells.keys())
                row_arr = [cells.get(i) for i in range(1, max_col + 1)]
                parsed_rows.append((row_num, row_arr))

        header_row = None
        header_num = None
        for row_num, arr in parsed_rows[:30]:
            non_empty = sum(1 for item in arr if self.clean_text(item) is not None)
            if non_empty >= 3:
                header_row = arr
                header_num = row_num
                break

        if header_row is None or header_num is None:
            raise HTTPException(status_code=400, detail="未能找到表头行")

        headers = [self.clean_text(h) for h in header_row]
        data_rows: list[dict[str, Any]] = []

        for row_num, arr in parsed_rows:
            if row_num <= header_num:
                continue
            if all(self.clean_text(item) is None for item in arr):
                continue

            row_dict: dict[str, Any] = {"__row_num__": row_num}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = arr[idx] if idx < len(arr) else None
                row_dict[header] = value
            data_rows.append(row_dict)

        missing = [
            column for column in self.REQUIRED_COLUMNS
            if column not in headers and column not in self.IGNORED_MISSING_FIELDS
        ]
        if missing:
            raise HTTPException(status_code=422, detail={"missing_columns": missing})

        return data_rows

    def build_row_context(self, raw: dict[str, Any]) -> dict[str, Any]:
        return {
            "row_num": raw.get("__row_num__"),
            "raw": raw,
            "record": {
                "case_id": self.to_int(raw.get("案件ID")),
                "debtor_name": self.clean_text(raw.get("姓名")),
                "uid": self.normalize_identifier(raw.get("UID")),
                "household_address": self.clean_text(raw.get("户籍地")),
                "application_code": self.normalize_identifier(raw.get("进件编码")),
                "transfer_notice_url": self.clean_text(raw.get("债转公告链接")),
                "transfer_notice_no": self.normalize_identifier(raw.get("债转公告序号")),
                "transfer_agreement_no": self.normalize_identifier(raw.get("债转协议序号")),
                "project_name": self.clean_text(raw.get("项目名称")),
                "province": self.clean_text(raw.get("省")),
                "city": self.clean_text(raw.get("市")),
                "entrust_date": self.parse_excel_serial_date(raw.get("委案日期")),
                "entrust_batch_no": self.clean_text(raw.get("委案批次号")),
                "entrust_org": self.clean_text(raw.get("委案机构")),
                "entrusted_principal_balance": self.to_decimal(raw.get("委案剩本金额")),
                "entrusted_total_claim": self.to_decimal(raw.get("委案债权总额")),
                "disposal_type": self.clean_text(raw.get("处置方式")),
                "case_followers": [],
            },
        }

    def serialize_row_detail(self, raw: dict[str, Any]) -> dict[str, Any]:
        serialized = {"row_num": raw.get("__row_num__")}
        for field_name in self.REQUIRED_COLUMNS:
            field = self.FIELD_MAPPING[field_name]
            serialized[field] = self.clean_text(raw.get(field_name))
            serialized[f"{field}_label"] = field_name
        return serialized

    def build_validation_issues(self, row_contexts: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        issues: list[dict[str, Any]] = []
        per_field: dict[str, dict[str, int]] = {}
        # 记录已为哪些行预生成了 ID，避免同一行多个缺失字段重复生成
        row_pregenerated_ids: dict[int, int] = {}

        for row_context in row_contexts:
            raw = row_context["raw"]
            row_num = row_context["row_num"]
            row_detail = self.serialize_row_detail(raw)
            for field_name in self.REQUIRED_COLUMNS:
                if field_name in self.IGNORED_MISSING_FIELDS:
                    continue

                value = raw.get(field_name)
                if self.clean_text(value) is not None:
                    continue

                # 该行首次出现缺失时，提前生成 ID 并写入 record
                if row_num not in row_pregenerated_ids:
                    pre_id = generate_snowflake_id()
                    row_pregenerated_ids[row_num] = pre_id
                    row_context["record"]["id"] = pre_id

                severity = "critical" if field_name in self.CRITICAL_FIELDS else "warning"
                issue = {
                    "row_num": row_num,
                    "record_id": str(row_pregenerated_ids[row_num]),
                    "field": self.FIELD_MAPPING[field_name],
                    "field_label": field_name,
                    "severity": severity,
                    "message": f"第 {row_num} 行缺少{field_name}",
                    "row_detail": row_detail,
                }
                issues.append(issue)

                stats = per_field.setdefault(field_name, {"critical": 0, "warning": 0})
                stats[severity] += 1

        summary = {
            "critical_count": sum(1 for item in issues if item["severity"] == "critical"),
            "warning_count": sum(1 for item in issues if item["severity"] == "warning"),
            "has_critical": any(item["severity"] == "critical" for item in issues),
            "fields": [
                {
                    "field": self.FIELD_MAPPING[field_label],
                    "field_label": field_label,
                    "critical_count": counts["critical"],
                    "warning_count": counts["warning"],
                }
                for field_label, counts in per_field.items()
            ],
        }
        return issues, summary

    @staticmethod
    def col_to_num(col: str) -> int:
        num = 0
        for ch in col:
            num = num * 26 + (ord(ch) - 64)
        return num

    @staticmethod
    def clean_text(value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text if text else None

    def normalize_identifier(self, value: Any) -> str | None:
        text = self.clean_text(value)
        if text is None:
            return None
        if text.endswith(".0"):
            int_part = text[:-2]
            if int_part.isdigit():
                return int_part
        return text

    def to_int(self, value: Any) -> int | None:
        text = self.normalize_identifier(value)
        if text is None:
            return None
        try:
            return int(text)
        except ValueError:
            return None

    def to_decimal(self, value: Any) -> str | None:
        text = self.clean_text(value)
        if text is None:
            return None
        try:
            return str(Decimal(text))
        except (InvalidOperation, ValueError):
            return None

    def parse_excel_serial_date(self, value: Any) -> str | None:
        text = self.clean_text(value)
        if text is None:
            return None

        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return dt.datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                pass

        try:
            serial = float(text)
            date_obj = dt.datetime(1899, 12, 30) + dt.timedelta(days=serial)
            return date_obj.date().isoformat()
        except ValueError:
            return None

    def get_cell_value(self, cell: ET.Element, shared: list[str]) -> str | None:
        cell_type = cell.attrib.get("t")
        v = cell.find("a:v", self.NS)

        if v is None:
            inline_str = cell.find("a:is", self.NS)
            if inline_str is None:
                return None
            pieces = [t.text or "" for t in inline_str.findall(".//a:t", self.NS)]
            text = "".join(pieces)
            return text if text else None

        raw = v.text
        if raw is None:
            return None

        if cell_type == "s":
            try:
                return shared[int(raw)]
            except (ValueError, IndexError):
                return raw

        return raw

    EXPORT_HEADERS = [
        ("案件ID", "case_id"), ("姓名", "debtor_name"), ("UID", "uid"),
        ("户籍地", "household_address"), ("进件编码", "application_code"),
        ("债转公告链接", "transfer_notice_url"), ("债转公告序号", "transfer_notice_no"),
        ("债转协议序号", "transfer_agreement_no"), ("项目名称", "project_name"),
        ("省", "province"), ("市", "city"), ("委案日期", "entrust_date"),
        ("委案批次号", "entrust_batch_no"), ("委案机构", "entrust_org"),
        ("委案剩本金额", "entrusted_principal_balance"), ("委案债权总额", "entrusted_total_claim"),
        ("处置方式", "disposal_type"),
    ]

    async def export_excel(self, file_hash: str, exclude_ids: list[str], order: str = "asc", sort_by: str = "case_status") -> tuple[io.BytesIO, str]:
        from collection_compiler_service.repository.CaseAssignmentDetailRepository import CaseAssignmentDetailRepository
        repo = CaseAssignmentDetailRepository(db_name="collection_compiler")
        async with repo as repository:
            records = await repository.find_by_hash_exclude_ids(file_hash, exclude_ids, order, sort_by)

        wb = openpyxl.Workbook()
        ws = wb.active

        from openpyxl.styles import PatternFill, Font, Alignment
        blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        gray_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
        white_bold = Font(bold=True, color="000000")
        dark_bold = Font(bold=True, color="000000")

        blue_headers = {"姓名", "户籍地", "省", "市"}
        yellow_headers = {"案件ID", "UID", "进件编码", "委案日期", "委案批次号", "委案机构", "委案剩本金额", "委案债权总额", "处置方式"}

        for col_idx, (header, _) in enumerate(self.EXPORT_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.alignment = Alignment(horizontal="center")
            if header in blue_headers:
                cell.fill = blue_fill
                cell.font = white_bold
            elif header in yellow_headers:
                cell.fill = yellow_fill
                cell.font = dark_bold
            else:
                cell.fill = gray_fill
                cell.font = dark_bold

        for r in records:
            ws.append([getattr(r, f) for _, f in self.EXPORT_HEADERS])

        # 列宽自适应
        for col_idx, (header, field) in enumerate(self.EXPORT_HEADERS, 1):
            max_len = len(header)
            for r in records:
                val = getattr(r, field)
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = quote(f"export_{file_hash[:8]}.xlsx")
        return buffer, filename


excel_upload_service = ExcelUploadService()